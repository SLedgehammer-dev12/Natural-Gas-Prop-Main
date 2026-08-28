"""
Report generation utilities.

Generates formatted text reports from calculation results with detailed logging.
"""

from datetime import datetime
from typing import Dict, Any, List, Tuple, Optional
from pathlib import Path
import logging
import os
import tempfile
from natural_gas_main.config.settings import config
class ReportGenerator:
    """Generates formatted text reports from calculation results."""

    @staticmethod
    def _setup_pdf_font(pdf: "FPDF") -> str:
        """
        Register a Unicode TrueType font for Turkish characters and symbols.

        FPDF core fonts such as Helvetica/Arial are Latin-1 only. Using them
        with text like "DOĞAL", "Özellik", "m³" or "ρ" raises an exception.
        """
        font_sets = []

        try:
            from matplotlib import font_manager

            regular = font_manager.findfont("DejaVu Sans", fallback_to_default=False)
            bold = font_manager.findfont("DejaVu Sans:style=normal:weight=bold", fallback_to_default=False)
            italic = font_manager.findfont("DejaVu Sans:style=oblique", fallback_to_default=False)
            font_sets.append(("DejaVu", regular, bold, italic))
        except Exception:
            pass

        # Bundled DejaVu fonts shipped inside the matplotlib package: these are
        # always present when matplotlib is installed (it is a hard dependency
        # of the UI), so the report does not depend on OS font paths.
        try:
            import matplotlib
            bundled_dir = os.path.join(os.path.dirname(matplotlib.__file__), "mpl-data", "fonts", "ttf")
            bundled_regular = os.path.join(bundled_dir, "DejaVuSans.ttf")
            if os.path.exists(bundled_regular):
                font_sets.append((
                    "DejaVu",
                    bundled_regular,
                    os.path.join(bundled_dir, "DejaVuSans-Bold.ttf"),
                    os.path.join(bundled_dir, "DejaVuSans-Oblique.ttf"),
                ))
        except Exception:
            pass

        font_sets.extend([
            (
                "ArialUnicode",
                r"C:\Windows\Fonts\arial.ttf",
                r"C:\Windows\Fonts\arialbd.ttf",
                r"C:\Windows\Fonts\ariali.ttf",
            ),
            (
                "SegoeUI",
                r"C:\Windows\Fonts\segoeui.ttf",
                r"C:\Windows\Fonts\segoeuib.ttf",
                r"C:\Windows\Fonts\segoeuii.ttf",
            ),
            (
                "HelveticaMac",
                "/System/Library/Fonts/Helvetica.dfont",
                "/System/Library/Fonts/Helvetica.dfont",
                "/System/Library/Fonts/Helvetica.dfont",
            ),
            (
                "ArialMac",
                "/Library/Fonts/Arial.ttf",
                "/Library/Fonts/Arial Bold.ttf",
                "/Library/Fonts/Arial Italic.ttf",
            ),
            (
                "DejaVuMac",
                "/Library/Fonts/DejaVuSans.ttf",
                "/Library/Fonts/DejaVuSans-Bold.ttf",
                "/Library/Fonts/DejaVuSans-Oblique.ttf",
            ),
            (
                "DejaVuLinux",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf",
            ),
            (
                "RobotoAndroid",
                "/system/fonts/Roboto-Regular.ttf",
                "/system/fonts/Roboto-Bold.ttf",
                "/system/fonts/Roboto-Italic.ttf",
            ),
            (
                "NotoSansAndroid",
                "/system/fonts/NotoSans-Regular.ttf",
                "/system/fonts/NotoSans-Bold.ttf",
                "/system/fonts/NotoSans-Italic.ttf",
            ),
            (
                "LiberationLinux",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf",
            ),
        ])

        for family, regular, bold, italic in font_sets:
            if not regular or not os.path.exists(regular):
                continue
            pdf.add_font(family, "", regular)
            pdf.add_font(family, "B", bold if bold and os.path.exists(bold) else regular)
            pdf.add_font(family, "I", italic if italic and os.path.exists(italic) else regular)
            return family

        raise RuntimeError(
            "PDF raporu için Unicode font bulunamadı. Android Roboto/NotoSans veya DejaVu Sans, Windows Arial/Segoe UI fontları gerekli."
        )
    
    @staticmethod
    def generate_text_report(
        input_params: Dict[str, Any],
        results: List[Tuple[str, str, str]],
        gas_composition: List[Tuple[str, float]],
        log_file: Optional[str] = None,
        include_full_log: bool = False,
        comparison_results: Optional[List[List[str]]] = None
    ) -> str:
        """
        Generate formatted text report with timestamped log.
        
        Args:
            input_params: Dictionary with calculation inputs
                - temperature: str (with unit)
                - pressure: str (with unit)
                - backend: str
                - volume: Optional[str] (with unit)
                - fraction_type: str ("molar" or "mass")
            results: List of (property, value, unit) tuples
            gas_composition: List of (gas_name, fraction) tuples
            log_file: Optional path to log file for extracting calculation logs
            include_full_log: If True, include all recent log entries
            
        Returns:
            Formatted report as string
        """
        report_lines = []
        timestamp = datetime.now()
        
        # Header
        report_lines.append("=" * 90)
        report_lines.append(" TERMODİNAMİK GAZ KARIŞIMI HESAPLAMA RAPORU")
        report_lines.append(" Natural Gas Prop Main - Modüler Sürüm")
        report_lines.append(f" Rapor Tarihi: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
        report_lines.append(f" Timestamp: {timestamp.isoformat()}")
        report_lines.append("=" * 90)
        report_lines.append("")
        
        # Input parameters
        report_lines.append("╔" + "═" * 88 + "╗")
        report_lines.append("║" + " GİRİLEN PARAMETRELER".center(88) + "║")
        report_lines.append("╚" + "═" * 88 + "╝")
        report_lines.append("")
        report_lines.append(f"  Sıcaklık           : {input_params.get('temperature', 'N/A')}")
        report_lines.append(f"  Basınç             : {input_params.get('pressure', 'N/A')}")
        report_lines.append(f"  Hesaplama Yöntemi  : {input_params.get('backend', 'N/A')}")
        
        if 'volume' in input_params and input_params['volume']:
            report_lines.append(f"  Hacim (ACM)        : {input_params['volume']} m³")
        
        fraction_type = input_params.get('fraction_type', 'molar')
        report_lines.append(f"  Kompozisyon Tipi   : {fraction_type.capitalize()}")
        report_lines.append("")
        
        # Gas composition
        report_lines.append("  Gaz Kompozisyonu:")
        report_lines.append("  " + "-" * 60)
        for gas_name, fraction in gas_composition:
            report_lines.append(f"    • {gas_name:<30s} : {fraction:>8.4f} %")
        report_lines.append("")
        
        # Results
        report_lines.append("╔" + "═" * 88 + "╗")
        report_lines.append("║" + " HESAPLAMA SONUÇLARI".center(88) + "║")
        report_lines.append("╚" + "═" * 88 + "╝")
        report_lines.append("")
        report_lines.append(f"  {'Özellik':<42} | {'Değer':<22} | {'Birim'}")
        report_lines.append("  " + "-" * 85)
        
        for prop, value, unit in results:
            if prop.startswith('-'):
                # Section header
                report_lines.append("")
                report_lines.append(f"  {prop}")
                report_lines.append("  " + "-" * 85)
            else:
                report_lines.append(f"  {prop:<42} | {value:<22} | {unit}")
        
        report_lines.append("")
        
        # Model comparison matrix
        if comparison_results and len(comparison_results) > 1:
            report_lines.append("╔" + "═" * 88 + "╗")
            report_lines.append("║" + " MODEL KARŞILAŞTIRMA MATRİSİ".center(88) + "║")
            report_lines.append("╚" + "═" * 88 + "╝")
            report_lines.append("")
            for row in comparison_results:
                report_lines.append("  " + "  |  ".join(str(c) for c in row))
            report_lines.append("")
        
        # Calculation Log Section
        report_lines.append("╔" + "═" * 88 + "╗")
        report_lines.append("║" + " HESAPLAMA LOG KAYDI (TIMESTAMP'Lİ)".center(88) + "║")
        report_lines.append("╚" + "═" * 88 + "╝")
        report_lines.append("")
        
        log_entries = ReportGenerator._get_calculation_log(log_file, include_full_log)
        
        if log_entries:
            report_lines.append("  Son Hesaplama İşlemlerinin Detaylı Kaydı:")
            report_lines.append("  " + "-" * 85)
            for entry in log_entries:
                report_lines.append(f"  {entry}")
        else:
            report_lines.append("  [Log kaydı bulunamadı veya okunamadı]")
        
        report_lines.append("")
        
        # Engineering disclaimer
        report_lines.append("╔" + "═" * 88 + "╗")
        report_lines.append("║" + " MÜHENDİSLİK SORUMLULUK REDDİ".center(88) + "║")
        report_lines.append("╚" + "═" * 88 + "╝")
        report_lines.append("")
        for line in ReportGenerator._wrap_text(config.ENGINEERING_DISCLAIMER, 84):
            report_lines.append("  " + line)
        report_lines.append("")
        
        # Footer
        report_lines.append("=" * 90)
        report_lines.append(f"  Rapor Oluşturma Zamanı: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report_lines.append("  Natural Gas Prop Main © 2025 Kompresör Pompa")
        report_lines.append("=" * 90)
        
        return "\n".join(report_lines)
    
    @staticmethod
    def _wrap_text(text: str, width: int) -> List[str]:
        """Word-wrap a long paragraph into lines of at most `width` chars."""
        words = text.split()
        lines: List[str] = []
        current = ""
        for word in words:
            if len(current) + len(word) + 1 > width:
                if current:
                    lines.append(current)
                current = word
            else:
                current = f"{current} {word}".strip()
        if current:
            lines.append(current)
        return lines

    @staticmethod
    def _get_calculation_log(log_file: Optional[str], include_full: bool = False) -> List[str]:
        """
        Extract calculation log entries from log file.
        
        Args:
            log_file: Path to log file
            include_full: If True, include all recent entries; if False, filter calculation-related
            
        Returns:
            List of formatted log entries
        """
        if not log_file or not os.path.exists(log_file):
            return []
        
        try:
            with open(log_file, 'r', encoding='utf-8-sig') as f:
                lines = f.readlines()
            
            # Find the start of the latest calculation
            calc_start_idx = 0
            for i in range(len(lines) - 1, -1, -1):
                if "Trying backend:" in lines[i] or "Hesaplanıyor" in lines[i]:
                    calc_start_idx = max(0, i - 2)
                    break
            
            # Get recent lines
            if calc_start_idx > 0:
                recent_lines = lines[calc_start_idx:]
            else:
                recent_lines = lines[-100:] if len(lines) > 100 else lines
            
            # Filter for calculation-related entries if not including full log
            if not include_full:
                keywords = [
                    'Trying backend',
                    'Successfully calculated',
                    'failed',
                    'HHV',
                    'LHV',
                    'Component-based',
                    'Reference',
                    'Creating state',
                    'Backend',
                    'Heating',
                    'calculation',
                    'NeqSim'
                ]
                filtered_lines = [
                    line.strip() for line in recent_lines
                    if any(keyword in line for keyword in keywords)
                ]
                return filtered_lines[-20:]  # Last 20 relevant entries
            else:
                return [line.strip() for line in recent_lines]
                
        except Exception as e:
            return [f"[Log okuma hatası: {e}]"]
    
    @staticmethod
    def generate_pdf_report(
        input_params: Dict[str, Any],
        results: List[Tuple[str, str, str]],
        gas_composition: List[Tuple[str, float]],
        file_path: str,
        plot_image_path: Optional[str] = None,
        comparison_results: Optional[List[List[str]]] = None
    ) -> None:
        """
        Generate professional PDF report using fpdf2.
        
        Args:
            input_params: Dictionary with calculation inputs
            results: List of (property, value, unit) tuples
            gas_composition: List of (gas_name, fraction) tuples
            file_path: Output PDF path
            plot_image_path: Optional path to phase envelope plot image
            comparison_results: Optional comparison matrix (header row first)
        """
        from fpdf import FPDF  # lazy import to avoid numpy dependency at startup

        pdf = FPDF()
        pdf.add_page()
        font_family = ReportGenerator._setup_pdf_font(pdf)
        
        pdf.set_font(font_family, "B", 16)
        
        # Header
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(0, 15, "DOĞAL GAZ ÖZELLİKLERİ HESAPLAMA RAPORU", new_x="LMARGIN", new_y="NEXT", align='C', fill=True)
        pdf.set_font(font_family, "", 10)
        pdf.cell(0, 8, f"Rapor Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", new_x="LMARGIN", new_y="NEXT", align='R')
        pdf.ln(5)
        
        # Section 1: Inputs
        pdf.set_font(font_family, "B", 12)
        pdf.set_text_color(31, 83, 141) # Dark Blue
        pdf.cell(0, 10, "1. GİRİLEN PARAMETRELER", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        pdf.set_font(font_family, "", 10)
        
        input_data = [
            ["Sıcaklık", input_params.get('temperature', 'N/A')],
            ["Basınç", input_params.get('pressure', 'N/A')],
            ["Hesaplama Yöntemi", input_params.get('backend', 'N/A')],
            ["Kompozisyon Tipi", input_params.get('fraction_type', 'molar').capitalize()]
        ]
        if 'volume' in input_params and input_params['volume']:
            input_data.append(["Hacim (ACM)", f"{input_params['volume']} m³"])
            
        for key, val in input_data:
            pdf.cell(50, 8, f"{key}:", border='B')
            pdf.cell(0, 8, str(val), border='B', new_x="LMARGIN", new_y="NEXT")
            
        pdf.ln(5)
        
        # Section 2: Composition Table
        pdf.set_font(font_family, "B", 12)
        pdf.set_text_color(31, 83, 141)
        pdf.cell(0, 10, "2. GAZ KOMPOZİSYONU", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        pdf.set_font(font_family, "B", 10)
        
        # Table Header
        pdf.cell(90, 8, "Bileşen", border=1, fill=True)
        pdf.cell(0, 8, "Oran (%)", border=1, fill=True, new_x="LMARGIN", new_y="NEXT", align='C')
        
        pdf.set_font(font_family, "", 10)
        for gas, frac in gas_composition:
            pdf.cell(90, 8, f" {gas}", border=1)
            pdf.cell(0, 8, f"{frac:>8.4f}", border=1, new_x="LMARGIN", new_y="NEXT", align='C')
            
        pdf.ln(10)
        
        # Add Plot if exists
        if plot_image_path and os.path.exists(plot_image_path):
            try:
                # Add a new page if not enough space
                if pdf.get_y() > 150:
                    pdf.add_page()
                pdf.image(plot_image_path, x=10, w=190)
                pdf.ln(5)
            except Exception as e:
                logging.error(f"Failed to add image to PDF: {e}")
        
        # Section 3: Results
        if pdf.get_y() > 200:
            pdf.add_page()
            
        pdf.set_font(font_family, "B", 12)
        pdf.set_text_color(31, 83, 141)
        pdf.cell(0, 10, "3. HESAPLAMA SONUÇLARI", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        
        # Results Table
        pdf.set_font(font_family, "B", 10)
        pdf.cell(100, 8, "Özellik", border=1, fill=True)
        pdf.cell(50, 8, "Değer", border=1, fill=True, align='C')
        pdf.cell(0, 8, "Birim", border=1, fill=True, new_x="LMARGIN", new_y="NEXT", align='C')
        
        pdf.set_font(font_family, "", 9)
        for prop, value, unit in results:
            if prop.startswith('-'):
                # Section header
                pdf.set_font(font_family, "I", 9)
                pdf.set_fill_color(245, 245, 245)
                pdf.cell(0, 7, prop.strip('- '), new_x="LMARGIN", new_y="NEXT", fill=True, border=1)
                pdf.set_font(font_family, "", 9)
            else:
                pdf.cell(100, 7, f" {prop}", border=1)
                pdf.cell(50, 7, value, border=1, align='C')
                pdf.cell(0, 7, unit, border=1, new_x="LMARGIN", new_y="NEXT", align='C')
        
        # Section 4: Model comparison matrix
        if comparison_results and len(comparison_results) > 1:
            if pdf.get_y() > 190:
                pdf.add_page()
            pdf.set_font(font_family, "B", 12)
            pdf.set_text_color(31, 83, 141)
            pdf.cell(0, 10, "4. MODEL KARŞILAŞTIRMA MATRİSİ", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            pdf.set_font(font_family, "B", 8)
            # header
            header = comparison_results[0]
            col_w = 150.0 / max(len(header), 1)
            for idx, h in enumerate(header):
                last = idx == len(header) - 1
                pdf.cell(col_w, 7, f" {h}", border=1, fill=True,
                         new_x="LMARGIN" if last else "RIGHT",
                         new_y="NEXT" if last else "LAST")
            pdf.set_font(font_family, "", 8)
            for row in comparison_results[1:]:
                for idx, val in enumerate(row):
                    last = idx == len(row) - 1
                    pdf.cell(col_w, 6, f" {val}", border=1,
                             new_x="LMARGIN" if last else "RIGHT",
                             new_y="NEXT" if last else "LAST")
                if pdf.get_y() > 265:
                    pdf.add_page()
                    pdf.set_font(font_family, "", 8)

        # Engineering disclaimer section
        if pdf.get_y() > 200:
            pdf.add_page()
        pdf.set_font(font_family, "B", 11)
        pdf.set_text_color(178, 34, 34)
        pdf.cell(0, 9, "MÜHENDİSLİK SORUMLULUK REDDİ", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(80, 80, 80)
        pdf.set_font(font_family, "", 8)
        for line in ReportGenerator._wrap_text(config.ENGINEERING_DISCLAIMER, 95):
            pdf.cell(0, 4.5, line, new_x="LMARGIN", new_y="NEXT")
        
        # Footer
        pdf.set_y(-20)
        pdf.set_font(font_family, "I", 8)
        y = pdf.get_y()
        pdf.cell(0, 10, "Natural Gas Prop Main © 2026 Kompresör Pompa", align='R')
        pdf.set_xy(pdf.l_margin, y)
        pdf.cell(0, 10, f"Sayfa {pdf.page_no()}", align='L')
        
        pdf.output(file_path)
    
    @staticmethod
    def export_csv(
        results: List[Tuple[str, str, str]],
        gas_composition: List[Tuple[str, float]],
        file_path: str,
        comparison_results: Optional[List[List[str]]] = None
    ) -> None:
        """
        Export results to a UTF-8 (BOM) CSV file that opens correctly in Excel.

        Args:
            results: List of (property, value, unit) tuples
            gas_composition: List of (gas_name, fraction) tuples
            file_path: Output CSV path
            comparison_results: Optional comparison matrix
        """
        import csv as _csv

        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = _csv.writer(f, delimiter=';')
            writer.writerow(["Gaz Kompozisyonu", "Oran (%)"])
            for gas, frac in gas_composition:
                writer.writerow([gas, f"{frac:.6f}"])
            writer.writerow([])
            writer.writerow(["Özellik", "Değer", "Birim"])
            for prop, value, unit in results:
                writer.writerow([prop, value, unit])
            if comparison_results and len(comparison_results) > 1:
                writer.writerow([])
                writer.writerow(["MODEL KARŞILAŞTIRMA MATRİSİ"])
                for row in comparison_results:
                    writer.writerow([str(c) for c in row])

    @staticmethod
    def export_excel(
        input_params: Dict[str, Any],
        results: List[Tuple[str, str, str]],
        gas_composition: List[Tuple[str, float]],
        file_path: str,
        comparison_results: Optional[List[List[str]]] = None
    ) -> None:
        """
        Export results to a formatted Excel (.xlsx) workbook using openpyxl.

        Args:
            input_params: Dictionary with calculation inputs
            results: List of (property, value, unit) tuples
            gas_composition: List of (gas_name, fraction) tuples
            file_path: Output XLSX path
            comparison_results: Optional comparison matrix
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as e:
            raise ImportError(
                "Excel (.xlsx) dışa aktarımı için 'openpyxl' paketi gerekli: "
                "pip install openpyxl"
            ) from e

        wb = Workbook()

        # Sheet 1: Results
        ws = wb.active
        ws.title = "Sonuçlar"
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")

        ws.append(["GİRİLEN PARAMETRELER"])
        for key, label in [("temperature", "Sıcaklık"), ("pressure", "Basınç"),
                           ("backend", "Hesaplama Yöntemi"), ("fraction_type", "Kompozisyon Tipi")]:
            if key in input_params:
                ws.append([label, input_params[key]])
        if input_params.get("volume"):
            ws.append(["Hacim (ACM)", input_params["volume"]])
        ws.append([])

        ws.append(["GAZ KOMPOZİSYONU"])
        ws.append(["Bileşen", "Oran (%)"])
        for cell in ws[ws.max_row]:
            cell.font = header_font_white
            cell.fill = header_fill
        for gas, frac in gas_composition:
            ws.append([gas, round(frac, 6)])
        ws.append([])

        ws.append(["ÖZELLİK", "DEĞER", "BİRİM"])
        for cell in ws[ws.max_row]:
            cell.font = header_font_white
            cell.fill = header_fill
        for prop, value, unit in results:
            ws.append([prop, value, unit])

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

        # Sheet 2: Comparison matrix
        if comparison_results and len(comparison_results) > 1:
            ws2 = wb.create_sheet("Karşılaştırma")
            for row in comparison_results:
                ws2.append([str(c) for c in row])
            for cell in ws2[1]:
                cell.font = header_font_white
                cell.fill = header_fill
            for col_cells in ws2.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws2.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

        # Sheet 3: Disclaimer
        ws3 = wb.create_sheet("Sorumluluk Reddi")
        ws3["A1"] = "MÜHENDİSLİK SORUMLULUK REDDİ"
        ws3["A1"].font = Font(bold=True, color="B22222")
        ws3["A2"] = config.ENGINEERING_DISCLAIMER
        ws3["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws3.column_dimensions["A"].width = 100
        ws3.row_dimensions[2].height = 120

        wb.save(file_path)

    @staticmethod
    def save_to_file(report_content: str, file_path: str) -> None:
        """
        Save report to file.
        
        Args:
            report_content: Report text content
            file_path: Path to save file
            
        Raises:
            IOError: If file cannot be written
        """
        try:
            with open(file_path, 'w', encoding='utf-8-sig') as f:
                f.write(report_content)
        except Exception as e:
            raise IOError(f"Rapor kaydedilemedi: {str(e)}")
    
    @staticmethod
    def generate_and_save(
        input_params: Dict[str, Any],
        results: List[Tuple[str, str, str]],
        gas_composition: List[Tuple[str, float]],
        file_path: str,
        log_file: Optional[str] = None
    ) -> None:
        """
        Generate report and save to file in one step.
        
        Args:
            input_params: Calculation inputs
            results: Calculation results
            gas_composition: Gas composition list
            file_path: Path to save file
            log_file: Optional log file path for including calculation logs
        """
        report = ReportGenerator.generate_text_report(
            input_params,
            results,
            gas_composition,
            log_file=log_file
        )
        try:
            ReportGenerator.save_to_file(report, file_path)
        except IOError:
            raise IOError(
                f"Rapor dosyaya kaydedilemedi: {file_path}. "
                f"Hata: Dosya yazılamıyor veya yol geçersiz."
            )

